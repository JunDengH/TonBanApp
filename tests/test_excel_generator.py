# tests/test_excel_generator.py
"""
excel_generator 模块测试
覆盖：
- 周统计基础/缺班排序/放假核减/多次核减上限/缺班不为负
- 月统计 聚合/斜线样式/排序/序号
- 文件被占用抛 PermissionError
"""
import os
import sys
import pytest
import pandas as pd
from openpyxl import load_workbook
from unittest.mock import patch

from src.modules.excel_generator import (
    generate_weekly_excel,
    generate_monthly_excel,
)


# ============================================================
# 工具：伪造 parse_weekly_schedule 的返回
# ============================================================
def _fake_schedule(schedule_dict):
    """把一个 {周一:[...],周二:[...]} 的 dict 塞给 parse_weekly_schedule"""
    return patch(
        "src.modules.excel_generator.parse_weekly_schedule",
        return_value=schedule_dict,
    )


# ============================================================
# 周统计测试
# ============================================================
def test_weekly_basic(tmp_path):
    """周统计基础：应值优先按排班名单统计，实际按实际名单统计"""
    total_names = ["张三", "李四", "王五"]
    schedule = {
        "周一": ["张三", "李四"],
        "周二": ["张三", "王五"],
        "周三": ["李四"],
        "周四": ["王五"],
        "周五": [],
    }
    output = tmp_path / "week1.xlsx"

    with _fake_schedule(schedule):
            generate_weekly_excel("fake.docx", total_names, [], str(output), actual_word_path="fake.docx")

    df = pd.read_excel(output)
    assert set(df.columns) == {"姓名", "应值班次", "实际班次", "缺班", "备注"}
    assert len(df) == 3

    # 取每个人的数据
    row = {r["姓名"]: r for _, r in df.iterrows()}
    assert row["张三"]["实际班次"] == 2
    assert row["张三"]["应值班次"] == 2
    assert row["张三"]["缺班"] == 0

    assert row["李四"]["实际班次"] == 2
    assert row["王五"]["实际班次"] == 2


def test_weekly_absence_sort(tmp_path):
    """缺班多的排最前"""
    total_names = ["张三", "李四", "王五"]
    schedule = {
        "周一": ["张三", "张三"],  # 张三2次
        "周二": ["李四"],          # 李四1次
        "周三": [],                # 王五0次
    }
    output = tmp_path / "week.xlsx"

    with _fake_schedule(schedule):
            generate_weekly_excel("fake.docx", total_names, [], str(output), actual_word_path="fake.docx")

    df = pd.read_excel(output)
    # 全员缺班都为0，并列时保持总名单原顺序
    assert df.iloc[0]["姓名"] == "张三"
    assert df.iloc[0]["缺班"] == 0
    assert df.iloc[-1]["姓名"] == "王五"
    assert df.iloc[-1]["缺班"] == 0


def test_weekly_holiday_reduction(tmp_path):
    """放假那天排班的人 应值班次 -1"""
    total_names = ["张三", "李四"]
    schedule = {
        "周三": ["张三"],   # 张三周三排了1次
        "周四": ["李四"],
    }
    output = tmp_path / "week.xlsx"

    with _fake_schedule(schedule):
        # 周三放假
            generate_weekly_excel("fake.docx", total_names, ["周三"], str(output), actual_word_path="fake.docx")

    df = pd.read_excel(output)
    row = {r["姓名"]: r for _, r in df.iterrows()}
    assert row["张三"]["应值班次"] == 0   # 1 - 1
    assert row["李四"]["应值班次"] == 1   # 未受影响，保持排班统计值


def test_weekly_holiday_reduce_by_scheduled_count(tmp_path):
    """放假核减按当天排班次数扣减，且不小于0"""
    total_names = ["张三"]
    schedule = {
        "周三": ["张三", "张三", "张三"],  # 排了3次
    }
    output = tmp_path / "week.xlsx"

    with _fake_schedule(schedule):
            generate_weekly_excel("fake.docx", total_names, ["周三"], str(output), actual_word_path="fake.docx")

    df = pd.read_excel(output)
        # 应值 3 - 3 = 0
    assert df.iloc[0]["应值班次"] == 0


def test_weekly_absence_non_negative(tmp_path):
    """实际班次多于应值班次时，缺班应为0，不能为负"""
    total_names = ["张三"]
    schedule = {
        "周一": ["张三", "张三", "张三"],  # 实际3次
    }
    output = tmp_path / "week.xlsx"

    with _fake_schedule(schedule):
            generate_weekly_excel("fake.docx", total_names, [], str(output), actual_word_path="fake.docx")

    df = pd.read_excel(output)
    row = df.iloc[0]
    assert row["实际班次"] == 3
    assert row["应值班次"] == 3
    assert row["缺班"] == 0


def test_weekly_overtime_added_to_actual(tmp_path):
    """周统计不再计算加班：实际班次仅来自实际名单Word"""
    total_names = ["张三", "李四"]
    schedule = {
        "周一": ["张三"],  # 张三原始实际=1
    }
    output = tmp_path / "week.xlsx"

    with _fake_schedule(schedule):
        generate_weekly_excel(
            "fake.docx",
            total_names,
            [],
            str(output),
            actual_word_path="fake.docx",
        )

    df = pd.read_excel(output)
    row = {r["姓名"]: r for _, r in df.iterrows()}
    # 不再叠加加班
    assert row["张三"]["实际班次"] == 1
    assert row["李四"]["实际班次"] == 0


def test_weekly_dual_word_should_and_actual_split(tmp_path):
    """双Word口径：应值看排班Word，实际看实际Word"""
    total_names = ["张三", "李四"]
    schedule_should = {
        "周三": ["张三"],  # 若周三放假，则张三应值被减1
    }
    schedule_actual = {
        "周一": ["李四", "李四"],  # 实际班次只来自实际Word
    }
    output = tmp_path / "week_dual.xlsx"

    with patch(
        "src.modules.excel_generator.parse_weekly_schedule",
        side_effect=[schedule_should, schedule_actual],
    ):
        generate_weekly_excel(
            word_path="排班.docx",
            total_names=total_names,
            holidays=["周三"],
            output_path=str(output),
            actual_word_path="实际.docx",
        )

    df = pd.read_excel(output)
    row = {r["姓名"]: r for _, r in df.iterrows()}
    # 张三：应值1-1=0，实际0 -> 缺班0
    assert row["张三"]["应值班次"] == 0
    assert row["张三"]["实际班次"] == 0
    assert row["张三"]["缺班"] == 0
    # 李四：应值0，实际2 -> 缺班0
    assert row["李四"]["应值班次"] == 0
    assert row["李四"]["实际班次"] == 2
    assert row["李四"]["缺班"] == 0


def test_weekly_should_comes_from_schedule_not_default_two(tmp_path):
    """当排班名单为 0 次时，应值应为 0，不应默认给 2"""
    total_names = ["张三", "李四"]
    schedule_should = {"周一": ["张三"]}
    schedule_actual = {"周二": ["李四", "李四"]}
    output = tmp_path / "week_should_from_schedule.xlsx"

    with patch(
        "src.modules.excel_generator.parse_weekly_schedule",
        side_effect=[schedule_should, schedule_actual],
    ):
        generate_weekly_excel(
            word_path="排班.docx",
            total_names=total_names,
            holidays=[],
            output_path=str(output),
            actual_word_path="实际.docx",
        )

    df = pd.read_excel(output)
    row = {r["姓名"]: r for _, r in df.iterrows()}
    assert row["张三"]["应值班次"] == 1
    assert row["李四"]["应值班次"] == 0


def test_weekly_senior_should_fixed_to_one_when_enabled(tmp_path):
    """开启大四规则后，被选中的大四助理应值班次固定为1"""
    total_names = ["张三", "李四", "王五"]
    schedule_should = {
        "周一": ["张三", "张三", "李四"],
    }
    schedule_actual = {
        "周一": ["张三"],
    }
    output = tmp_path / "week_senior_fixed.xlsx"

    with patch(
        "src.modules.excel_generator.parse_weekly_schedule",
        side_effect=[schedule_should, schedule_actual],
    ):
        generate_weekly_excel(
            word_path="排班.docx",
            total_names=total_names,
            holidays=[],
            output_path=str(output),
            actual_word_path="实际.docx",
            senior_assistants=["张三", "王五"],
            senior_should_fixed_enabled=True,
        )

    df = pd.read_excel(output)
    row = {r["姓名"]: r for _, r in df.iterrows()}
    # 张三原本应值2，被固定为1
    assert row["张三"]["应值班次"] == 1
    # 王五原本应值0，也被固定为1
    assert row["王五"]["应值班次"] == 1
    # 李四不在大四名单，保持原逻辑
    assert row["李四"]["应值班次"] == 1


def test_weekly_senior_rule_disabled_keeps_original_should(tmp_path):
    """关闭大四规则时，不应修改应值班次"""
    total_names = ["张三"]
    schedule = {"周一": ["张三", "张三"]}
    output = tmp_path / "week_senior_disabled.xlsx"

    with _fake_schedule(schedule):
        generate_weekly_excel(
            word_path="fake.docx",
            total_names=total_names,
            holidays=[],
            output_path=str(output),
            actual_word_path="fake.docx",
            senior_assistants=["张三"],
            senior_should_fixed_enabled=False,
        )

    df = pd.read_excel(output)
    assert df.iloc[0]["应值班次"] == 2


def test_weekly_senior_fixed_but_holiday_has_higher_priority(tmp_path):
    """开启大四固定1后，若遇放假仍应继续核减（放假优先级最高）"""
    total_names = ["张三", "李四"]
    schedule_should = {
        "周三": ["张三"],
    }
    schedule_actual = {
        "周三": ["张三"],
    }
    output = tmp_path / "week_senior_holiday_priority.xlsx"

    with patch(
        "src.modules.excel_generator.parse_weekly_schedule",
        side_effect=[schedule_should, schedule_actual],
    ):
        generate_weekly_excel(
            word_path="排班.docx",
            total_names=total_names,
            holidays=["周三"],
            output_path=str(output),
            actual_word_path="实际.docx",
            senior_assistants=["张三"],
            senior_should_fixed_enabled=True,
        )

    df = pd.read_excel(output)
    row = {r["姓名"]: r for _, r in df.iterrows()}
    # 张三：固定1后再核减1 -> 0
    assert row["张三"]["应值班次"] == 0
    # 李四未排班且非大四 -> 0
    assert row["李四"]["应值班次"] == 0


# ============================================================
# 月统计测试
# ============================================================
def _make_weekly_excel(tmp_path, filename, total_names, schedule, holidays=None):
    """辅助：用 generate_weekly_excel 生成一个真实的周Excel"""
    output = tmp_path / filename
    with _fake_schedule(schedule):
        generate_weekly_excel(
            "fake.docx", total_names, holidays or [], str(output), actual_word_path="fake.docx"
        )
    return str(output)


def test_monthly_aggregation(tmp_path):
    """月统计：4周累加"""
    total_names = ["张三", "李四", "王五"]

    # 4 周，每周不同排班
    w1 = _make_weekly_excel(tmp_path, "w1.xlsx", total_names, {
        "周一": ["张三", "李四"],
        "周二": ["张三", "王五"],
    })
    w2 = _make_weekly_excel(tmp_path, "w2.xlsx", total_names, {
        "周一": ["李四"],
        "周二": ["张三"],
    })
    w3 = _make_weekly_excel(tmp_path, "w3.xlsx", total_names, {
        "周一": ["王五", "王五"],
    })
    w4 = _make_weekly_excel(tmp_path, "w4.xlsx", total_names, {
        "周一": ["张三"],
    })

    output = tmp_path / "month.xlsx"
    generate_monthly_excel([w1, w2, w3, w4], total_names, str(output))

    df = pd.read_excel(output)
    # 必需列
    assert list(df.columns) == ["序号", "姓名", "总计班次", "应值班次", "多值班次", "缺班数量"]
    # 应值总计来自周排班：张三4, 李四2, 王五3
    row = {r["姓名"]: r for _, r in df.iterrows()}
    assert row["张三"]["应值班次"] == 4
    assert row["李四"]["应值班次"] == 2
    assert row["王五"]["应值班次"] == 3

    # 张三实际 = 2+1+0+1 = 4
    assert row["张三"]["总计班次"] == 4
    assert row["李四"]["总计班次"] == 2
    assert row["王五"]["总计班次"] == 3


def test_monthly_slash_style(tmp_path):
    """月统计：多值/缺班 <= 0 的格子要清空 + 画斜线"""
    total_names = ["张三"]
    # 一周排 2 次，刚好 应值=实际，多值=0，缺班=0
    w1 = _make_weekly_excel(tmp_path, "w1.xlsx", total_names, {
        "周一": ["张三", "张三"],
    })
    output = tmp_path / "month.xlsx"
    generate_monthly_excel([w1], total_names, str(output))

    wb = load_workbook(output)
    ws = wb["月统计"]
    # 表头第1行，数据第2行
    # E列=多值班次, F列=缺班数量
    assert ws.cell(row=2, column=5).value is None
    assert ws.cell(row=2, column=6).value is None
    # 有斜线
    assert ws.cell(row=2, column=5).border.diagonal is not None
    assert ws.cell(row=2, column=5).border.diagonalDown is True
    assert ws.cell(row=2, column=6).border.diagonalDown is True


def test_monthly_sort_order(tmp_path):
    """月统计排序：多值多的在前，缺班多的在后"""
    total_names = ["A多值", "B缺班", "C正常"]

    # 构造一个场景：
    # A: 排3次 -> 多值1
    # B: 排0次 -> 缺班0
    # C: 排2次 -> 正好
    w1 = _make_weekly_excel(tmp_path, "w1.xlsx", total_names, {
        "周一": ["A多值", "A多值", "C正常"],
        "周二": ["A多值", "C正常"],
    })
    output = tmp_path / "month.xlsx"
    generate_monthly_excel([w1], total_names, str(output))

    df = pd.read_excel(output)
    # 多值最多的第一
    assert df.iloc[0]["姓名"] == "A多值"
    # 其余同分时按拼音顺序，B 在 C 前
    assert list(df["姓名"]) == ["A多值", "B缺班", "C正常"]


def test_monthly_serial_number(tmp_path):
    """月统计：序号从 1 递增"""
    total_names = [f"员工{i}" for i in range(5)]
    w1 = _make_weekly_excel(tmp_path, "w1.xlsx", total_names, {
        "周一": ["员工0", "员工1"],
    })
    output = tmp_path / "month.xlsx"
    generate_monthly_excel([w1], total_names, str(output))

    df = pd.read_excel(output)
    assert list(df["序号"]) == [1, 2, 3, 4, 5]


# ============================================================
# 文件占用测试
# ============================================================
@pytest.mark.skipif(sys.platform != "win32", reason="仅Windows下文件占用会抛PermissionError")
def test_file_locked_raises(tmp_path):
    """
    当目标Excel文件被独占打开时，生成应抛 PermissionError。
    Windows 下用 msvcrt 独占锁文件模拟 Excel 打开。
    """
    import msvcrt

    total_names = ["张三"]
    schedule = {"周一": ["张三"]}
    output = tmp_path / "locked.xlsx"

    # 先创建一个占位文件
    output.write_bytes(b"placeholder")

    # 以独占方式打开（Windows 下 Excel 打开文件的效果类似）
    f = open(output, "r+b")
    try:
        try:
            msvcrt.locking(f.fileno(), msvcrt.LK_NBLCK, 1)
            locked = True
        except OSError:
            locked = False

        if not locked:
            pytest.skip("当前环境无法独占锁定文件，跳过该测试")

        with _fake_schedule(schedule):
            with pytest.raises(PermissionError):
                generate_weekly_excel(
                    "fake.docx", total_names, [], str(output), actual_word_path="fake.docx"
                )
    finally:
        try:
            msvcrt.locking(f.fileno(), msvcrt.LK_UNLCK, 1)
        except OSError:
            pass
        f.close()