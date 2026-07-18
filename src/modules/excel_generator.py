# src/modules/excel_generator.py
"""
Excel生成模块：
A. 周统计Excel（5列）
"""
import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Side, Font, PatternFill
from openpyxl.utils import get_column_letter

from src.modules.word_parser import (
    parse_weekly_schedule,
    count_actual_shifts,
    find_typo_suspects,
)
from src.modules.schedule_parser import parse_schedule_file, scan_schedule_file


LONG_TERM_LEAVE_SHOULD = 2
WEEKLY_SHOULD_MAX = 2
SUPPORTED_HOLIDAYS = {"周一", "周二", "周三", "周四", "周五"}
SENIOR_MODE_NORMAL = "normal"
SENIOR_MODE_REDUCED = "reduced"
SENIOR_MODE_NONE = "none"
SENIOR_SHOULD_MODES = {SENIOR_MODE_NORMAL, SENIOR_MODE_REDUCED, SENIOR_MODE_NONE}


# ============================================================
# 子功能 A：周统计 Excel
# ============================================================
def generate_weekly_excel(
    word_path: str,
    total_names: list,
    holidays: list,
    output_path: str,
    actual_word_path: str | None = None,
    senior_assistants: list | None = None,
    senior_should_fixed_enabled: bool | str = False,
    long_term_leave_assistants: list | None = None,
    final_week_enabled: bool = False,
    corrections: dict | None = None,
) -> str:
    """
    生成周统计Excel。
    参数:
        word_path: 排班名单Word文件路径（用于应值班次计算）
        total_names: 总名单（来自 data_manager）
        holidays: 放假的星期列表，如 ["周三", "周五"]
        output_path: 输出文件完整路径（.xlsx）
        actual_word_path: 实际周统计人员名单Word路径（用于实际班次计算）；
                 为空时回退使用 word_path（兼容旧流程）
        senior_assistants: 被纳入毕业季特殊逻辑名单的姓名列表
        senior_should_fixed_enabled: 少值班模式下，毕业季助理应值班次强制为 1
        long_term_leave_assistants: 被标记为长期请假的助理；其应值班次视作 2，
                 但仍会被毕业季特殊逻辑覆盖
        final_week_enabled: 期末周规则。开启后排版表/值班表中每位助理仅出现一次，
                 但每次出现计为两次值班（应值/实际计数及放假核减均 ×2）；固定额度
                 （长期请假应值、毕业季模式）不翻倍。
        corrections: 暂时纠错映射 {"张灏琛": "张颢琛"}，将错字当作正确姓名计数，不修改原文件。
    返回:
        实际写入的文件路径
    """
    rows = build_weekly_rows(
        word_path=word_path,
        total_names=total_names,
        holidays=holidays,
        actual_word_path=actual_word_path,
        senior_assistants=senior_assistants,
        senior_should_fixed_enabled=senior_should_fixed_enabled,
        long_term_leave_assistants=long_term_leave_assistants,
        final_week_enabled=final_week_enabled,
        corrections=corrections,
    )
    return generate_weekly_excel_from_rows(rows, output_path)


def build_weekly_rows(
    word_path: str,
    total_names: list,
    holidays: list,
    actual_word_path: str | None = None,
    senior_assistants: list | None = None,
    senior_should_fixed_enabled: bool | str = False,
    long_term_leave_assistants: list | None = None,
    final_week_enabled: bool = False,
    corrections: dict | None = None,
) -> list[dict]:
    """计算周统计预览行，不写入文件。

    corrections: 暂时纠错映射，将排班/实际文件中的错字当作正确姓名计数。
    """
    schedule_for_should = parse_schedule_file(word_path, total_names, corrections)
    actual_source_path = actual_word_path or word_path
    schedule_for_actual = parse_weekly_schedule(actual_source_path, total_names, corrections)
    return build_weekly_rows_from_schedules(
        schedule_for_should=schedule_for_should,
        schedule_for_actual=schedule_for_actual,
        total_names=total_names,
        holidays=holidays,
        senior_assistants=senior_assistants,
        senior_should_fixed_enabled=senior_should_fixed_enabled,
        long_term_leave_assistants=long_term_leave_assistants,
        final_week_enabled=final_week_enabled,
    )


def build_weekly_rows_from_schedules(
    schedule_for_should: dict[str, list[str]],
    schedule_for_actual: dict[str, list[str]],
    total_names: list,
    holidays: list,
    senior_assistants: list | None = None,
    senior_should_fixed_enabled: bool | str = False,
    long_term_leave_assistants: list | None = None,
    final_week_enabled: bool = False,
) -> list[dict]:
    """计算已解析排班与实际名单的周统计行，不读取或写入文件。"""
    invalid_holidays = [holiday for holiday in (holidays or []) if holiday not in SUPPORTED_HOLIDAYS]
    if invalid_holidays:
        raise ValueError("放假日期只支持周一到周五。")

    # 期末周规则：每次出现计为两次值班，应值/实际计数及放假核减均 ×2。
    multiplier = 2 if final_week_enabled else 1

    # 0) 期末周与毕业季"少值班"冲突校验：期末周实际班次永远是偶数，而少值班应值=1，
    #    无法对齐，故在生成入口拦截，提示用户改配置。
    senior_should_mode = _normalize_senior_should_mode(senior_should_fixed_enabled)
    if (
        final_week_enabled
        and senior_should_mode == SENIOR_MODE_REDUCED
        and any(name in set(senior_assistants or []) for name in total_names)
    ):
        raise ValueError(
            "期末周规则与毕业季『少值班』模式冲突：期末周每次值班按 2 次计，"
            "『少值班』额度为 1 次无法对齐。请改为『无需值班』或『正常值班』，"
            "或关闭期末周规则后再生成。"
        )

    actual = {n: c * multiplier for n, c in count_actual_shifts(schedule_for_actual, total_names).items()}

    # 3) 应值班次：优先按排班名单统计（×multiplier），再按放假日期核减
    should = {n: c * multiplier for n, c in count_actual_shifts(schedule_for_should, total_names).items()}
    holiday_reduction_counter = {n: 0 for n in total_names}
    for holiday in holidays or []:
        if holiday not in schedule_for_should:
            continue
        # 统计当日每人排班次数
        day_names = schedule_for_should[holiday]
        per_person = {}
        for n in day_names:
            per_person[n] = per_person.get(n, 0) + 1
        # 核减：期末周下每次出现按 2 次核减；holiday_reduction_counter 仍记原始次数，
        # 仅供毕业季"少值班"分支使用（该分支在期末周下已被步骤 0 拦截）。
        for person, times in per_person.items():
            if person in holiday_reduction_counter:
                holiday_reduction_counter[person] += times
            should[person] = max(0, should[person] - times * multiplier)

    # 4) 长期请假规则：只对用户显式选择的助理生效。
    total_name_set = set(total_names)
    long_term_leave_names = {
        name for name in (long_term_leave_assistants or [])
        if name in total_name_set
    }
    for name in long_term_leave_names:
        should[name] = LONG_TERM_LEAVE_SHOULD

    # 5) 毕业季特殊逻辑：三档模式覆盖默认应值班次，并约束长期请假助理。
    #    （senior_should_mode 已在步骤 0 计算）
    if senior_should_mode != SENIOR_MODE_NORMAL and senior_assistants:
        senior_set = set(senior_assistants)
        for name in total_names:
            if name in senior_set:
                # 放假优先级更高：先固定为1，再按放假核减
                if senior_should_mode == SENIOR_MODE_NONE:
                    should[name] = 0
                else:
                    should[name] = max(0, 1 - holiday_reduction_counter.get(name, 0))

    rows = []
    for name in total_names:
        s = should[name]
        a = actual.get(name, 0)
        absence = max(0, s - a)
        rows.append({
            "姓名": name,
            "应值班次": s,
            "实际班次": a,
            "缺班": absence,
            "备注": "",
        })

    rows.sort(key=lambda row: -row["缺班"])
    return rows


def _weekly_word_sources(schedule_word_path, actual_word_path) -> list[tuple]:
    """
    返回去重后的 (来源标签, 路径) 列表，供周统计扫描复用。
    - 仅保留存在的路径；
    - 同一个文件被同时选为排班与实际时只保留排班来源，避免重复扫描 / 重复计数。
    """
    sources = []
    schedule_label = "排班 PDF" if str(schedule_word_path or "").lower().endswith(".pdf") else "排班 Word"
    for label, path in ((schedule_label, schedule_word_path), ("实际 Word", actual_word_path)):
        if not path or not os.path.exists(path):
            continue
        if any(os.path.samefile(path, kept) for _, kept in sources):
            continue
        sources.append((label, path))
    return sources


def collect_weekly_warnings(
    rows: list[dict],
    schedule_word_path: str,
    actual_word_path: str | None,
    total_names: list,
    corrections: dict | None = None,
) -> dict:
    """
    周统计生成前的软警告（不阻断生成）。

    corrections: 暂时纠错映射；已纠错的姓名不会出现在 unknown_names / typo_messages 中。

    返回:
        {
            "messages": [str, ...],            # 人类可读的提醒，逐条列出
            "highlight_names": set[str],       # 需要在预览里高亮的姓名（应值班次 > 2）
            "typo_messages_strong": [str,...], # 疑似打错字（高度疑似 + 疑似），单独弹窗用
            "typo_messages_weak": [str, ...],  # 形近待确认（读音不同），单独列出
            "typo_suspects": [dict, ...],      # 结构化明细，含 source / name / candidates / level
        }
    """
    messages = []
    typo_messages_strong = []
    typo_messages_weak = []
    typo_suspects = []
    highlight_names = {
        row["姓名"] for row in rows if row.get("应值班次", 0) > WEEKLY_SHOULD_MAX
    }
    if highlight_names:
        names_text = "、".join(sorted(highlight_names))
        messages.append(f"以下助理本周应值班次大于 {WEEKLY_SHOULD_MAX}：{names_text}")

    for label, path in _weekly_word_sources(schedule_word_path, actual_word_path):
        scan = scan_schedule_file(path, total_names, corrections)
        if scan["matched_count"] == 0:
            messages.append(f"{label} 未识别到任何总名单内姓名，请确认是否选错文件。")

        # 先从未知名里挑出疑似打错字，按档位单独、醒目地提醒；其余未知名仍按“已忽略”提示。
        suspects = find_typo_suspects(scan["unknown_names"], total_names)
        suspect_names = set()
        for suspect in suspects:
            suspect_names.add(suspect["name"])
            candidate_text = "、".join(suspect["candidates"])
            if suspect["level"] == "weak":
                typo_messages_weak.append(
                    f"{label} 中『{suspect['name']}』与总名单中『{candidate_text}』"
                    f"字形相近（读音不同），请确认是否同一人"
                )
            else:
                typo_messages_strong.append(
                    f"{label} 中『{suspect['name']}』与总名单中『{candidate_text}』"
                    f"高度相似，疑似打错字"
                )
            typo_suspects.append({
                "source": label,
                "name": suspect["name"],
                "candidates": suspect["candidates"],
                "level": suspect["level"],
            })

        remaining_unknown = [n for n in scan["unknown_names"] if n not in suspect_names]
        if remaining_unknown:
            unknown_text = "、".join(remaining_unknown[:10])
            if len(remaining_unknown) > 10:
                unknown_text += "…"
            messages.append(
                f"{label} 中以下内容不在总名单（已忽略，也可能是说明文字）：{unknown_text}"
            )

    return {
        "messages": messages,
        "highlight_names": highlight_names,
        "typo_messages_strong": typo_messages_strong,
        "typo_messages_weak": typo_messages_weak,
        "typo_suspects": typo_suspects,
    }


def collect_typo_suspects(
    schedule_word_path: str | None,
    actual_word_path: str | None,
    total_names: list,
) -> list[dict]:
    """
    扫描排班 / 实际 Word，返回疑似打错字明细，供生成前检查面板实时显示。

    与 collect_weekly_warnings 中的 typo 逻辑一致，但不依赖预览行、只做轻量扫描，
    路径不存在或与排班同一文件时静默跳过。返回:
        [{"source": str, "name": str, "candidates": [str, ...], "level": str}, ...]
    """
    results = []
    for label, path in _weekly_word_sources(schedule_word_path, actual_word_path):
        scan = scan_schedule_file(path, total_names)
        for suspect in find_typo_suspects(scan["unknown_names"], total_names):
            results.append({
                "source": label,
                "name": suspect["name"],
                "candidates": suspect["candidates"],
                "level": suspect["level"],
            })
    return results


def generate_weekly_excel_from_rows(rows: list[dict], output_path: str) -> str:
    """根据预览确认后的周统计行写入 Excel。"""
    normalized_rows = [_normalize_weekly_row(row) for row in rows]
    df = pd.DataFrame(normalized_rows, columns=["姓名", "应值班次", "实际班次", "缺班", "备注"])
    df = df.sort_values(by="缺班", ascending=False, kind="mergesort").reset_index(drop=True)

    _safe_write_excel(df, output_path, sheet_name="周统计", apply_weekly_style=True)
    return output_path


def _normalize_weekly_row(row: dict) -> dict:
    name = str(row.get("姓名", "")).strip()
    if not name:
        raise ValueError("周统计数据中存在空姓名。")
    should = _coerce_non_negative_int(row.get("应值班次", 0), name, "应值班次")
    actual = _coerce_non_negative_int(row.get("实际班次", 0), name, "实际班次")
    return {
        "姓名": name,
        "应值班次": should,
        "实际班次": actual,
        "缺班": max(0, should - actual),
        "备注": str(row.get("备注", "") or "").strip(),
    }


def _coerce_non_negative_int(value, name: str, column: str) -> int:
    if value is None:
        return 0
    if isinstance(value, str):
        text = value.strip()
        if not text:
            return 0
        if text.isdigit():
            return int(text)
        raise ValueError(f"{name} 的【{column}】必须为非负整数。")
    if isinstance(value, bool):
        raise ValueError(f"{name} 的【{column}】必须为非负整数。")
    try:
        number = float(value)
    except (TypeError, ValueError):
        raise ValueError(f"{name} 的【{column}】必须为非负整数。")
    if number < 0 or not number.is_integer():
        raise ValueError(f"{name} 的【{column}】必须为非负整数。")
    return int(number)


# ============================================================
# 公共：聚合 4 周 Excel
# ============================================================
def _normalize_senior_should_mode(value) -> str:
    if value in SENIOR_SHOULD_MODES:
        return value
    if value is True:
        return SENIOR_MODE_REDUCED
    return SENIOR_MODE_NORMAL


def aggregate_weekly_files(weekly_excel_paths: list, total_names: list):
    """
    读取多个周统计 Excel，按总名单聚合"实际班次""应值班次"。
    返回:
        (total_actual: dict[name,int], total_should: dict[name,int])
    说明:
        该函数同时为 excel_generator 与 word_generator 所共用，集中周表解析逻辑。
    """
    if len(weekly_excel_paths) < 1:
        raise ValueError("至少需要提供一个周统计Excel文件")

    total_actual = {n: 0 for n in total_names}
    total_should = {n: 0 for n in total_names}

    for path in weekly_excel_paths:
        df = pd.read_excel(path)
        required = {"姓名", "应值班次", "实际班次"}
        if not required.issubset(df.columns):
            raise ValueError(f"文件缺少必要列：{path}")
        for _, row in df.iterrows():
            name = str(row["姓名"]).strip()
            if name in total_actual:
                total_actual[name] += _read_non_negative_int(row["实际班次"], path, name, "实际班次")
                total_should[name] += _read_non_negative_int(row["应值班次"], path, name, "应值班次")

    return total_actual, total_should


def _read_non_negative_int(value, path, name: str, column: str) -> int:
    """Read shift counts from user-editable weekly Excel cells."""
    if pd.isna(value):
        return 0
    if isinstance(value, str):
        text = value.strip()
        if not text:
            return 0
        if text.isdigit():
            return int(text)
        raise ValueError(f"{path} 中 {name} 的【{column}】必须为非负整数。")
    if isinstance(value, bool):
        raise ValueError(f"{path} 中 {name} 的【{column}】必须为非负整数。")
    try:
        number = float(value)
    except (TypeError, ValueError):
        raise ValueError(f"{path} 中 {name} 的【{column}】必须为非负整数。")
    if number < 0 or not number.is_integer():
        raise ValueError(f"{path} 中 {name} 的【{column}】必须为非负整数。")
    return int(number)


# ============================================================
# 写入 + 样式
# ============================================================
def _safe_write_excel(df: pd.DataFrame, output_path: str,
                      sheet_name="Sheet1",
                      apply_weekly_style=False):
    """写入Excel并捕获文件被占用的情况"""
    os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)
    try:
        with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
            df.to_excel(writer, index=False, sheet_name=sheet_name)
    except PermissionError:
        raise PermissionError(
            f"无法写入 {output_path}，请先关闭已打开的该Excel文件后重试。"
        )

    # 应用样式
    wb = load_workbook(output_path)
    ws = wb[sheet_name]
    _apply_base_style(ws)

    wb.save(output_path)


def _apply_base_style(ws):
    """通用样式：居中、边框、列宽自适应、首行加粗"""
    thin = Side(border_style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center")

    max_row = ws.max_row
    max_col = ws.max_column

    # 列宽：按表头字符估算
    for col in range(1, max_col + 1):
        letter = get_column_letter(col)
        max_len = 0
        for row in range(1, max_row + 1):
            v = ws.cell(row=row, column=col).value
            if v is not None:
                max_len = max(max_len, len(str(v)))
        ws.column_dimensions[letter].width = max(10, max_len * 2 + 2)

    # 单元格格式
    for row in range(1, max_row + 1):
        for col in range(1, max_col + 1):
            cell = ws.cell(row=row, column=col)
            cell.border = border
            cell.alignment = center
            if row == 1:
                cell.font = Font(bold=True)
                cell.fill = PatternFill("solid", fgColor="D9E1F2")
