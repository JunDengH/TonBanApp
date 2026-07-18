"""
Excel 通讯录解析模块。

从一份类似「2026年资助中心助理通讯录.xlsx」的通讯录中自动识别每位助理的
姓名、部门、专业、年级，供总名单导入使用。

通讯录约定（容错解析，不依赖固定列顺序）：
- 含「姓名」和「专业年级」（或「专业」「年级」）的那一行视为表头行。
- 「部门」是合并单元格风格：仅每个部门块首行有值，后续行为空，需向下填充。
- 「姓名」可能含全角空格（如 张　爽），按汉字间全角空格归一化为 张爽。
- 「专业年级」常见 专业+届号（新闻22级），也兼容反序（24级法学）、带空格（材料 24级）、
  缩写（AIA22级）；届号 token 统一识别为 \\d{2,4}级。
"""
import re

from openpyxl import load_workbook

from src.modules.word_parser import NAME_REGEX


# 表头别名：用于在不同通讯录里定位列
_HEADER_ALIASES = {
    "department": ("部门", "部门名称", "所属部门"),
    "name": ("姓名", "名字"),
    "major_grade": ("专业年级", "专业及年级", "专业/年级"),
    "major": ("专业",),
    "grade": ("年级",),
}

_GRADE_TOKEN = re.compile(r"\d{2,4}\s*级")
_COHORT_NUMBER = re.compile(r"\d{2,4}")


def parse_contact_xlsx(xlsx_path: str) -> list[dict]:
    """解析 Excel 通讯录，返回助理记录列表。

    返回:
        [{"name": str, "department": str, "major": str, "grade": str}, ...]
        按出现顺序、姓名去重（首次出现为准）。
    异常:
        ValueError: 找不到含「姓名」「专业年级/专业」的表头行时抛出。
    """
    wb = load_workbook(xlsx_path, data_only=True)
    for ws in wb.worksheets:
        header = _locate_header(ws)
        if header is not None:
            return _parse_sheet(ws, header)
    raise ValueError("通讯录中未找到包含「姓名」和「专业年级」的表头行。")


def _locate_header(ws):
    """在工作表中定位表头行，返回 {字段: 列索引(0基)}；找不到返回 None。"""
    max_scan = min(ws.max_row, 20)
    for row_idx in range(1, max_scan + 1):
        cells = [_clean(c.value) for c in ws[row_idx]]
        col_map = {}
        for field, aliases in _HEADER_ALIASES.items():
            for col_idx, text in enumerate(cells):
                if text in aliases:
                    col_map[field] = col_idx
                    break
        has_name = "name" in col_map
        has_major = "major_grade" in col_map or "major" in col_map
        if has_name and has_major:
            col_map["_row"] = row_idx
            return col_map
    return None


def _parse_sheet(ws, header) -> list[dict]:
    name_col = header["name"]
    dept_col = header.get("department")
    major_grade_col = header.get("major_grade")
    major_col = header.get("major")
    grade_col = header.get("grade")
    start_row = header["_row"] + 1

    records = []
    seen = set()
    current_department = ""
    for row_idx in range(start_row, ws.max_row + 1):
        row = ws[row_idx]

        # 部门向下填充：本行有部门则更新，否则沿用上一非空部门。
        if dept_col is not None:
            dept_text = _clean(_cell(row, dept_col))
            if dept_text:
                current_department = dept_text

        name = _normalize_name(_cell(row, name_col))
        if not NAME_REGEX.fullmatch(name):
            continue
        if name in seen:
            continue
        seen.add(name)

        if major_grade_col is not None:
            major, grade = _split_major_grade(_cell(row, major_grade_col))
        else:
            major = _clean(_cell(row, major_col)) if major_col is not None else ""
            grade_raw = _clean(_cell(row, grade_col)) if grade_col is not None else ""
            _, grade = _split_major_grade(grade_raw)
            grade = grade or grade_raw

        records.append({
            "name": name,
            "department": current_department,
            "major": major,
            "grade": grade,
        })
    return records


def compute_graduation_seniors(records: list[dict]) -> list[str]:
    """识别毕业季助理：取届号最小（最靠前）的一组姓名。

    始终执行（不区分年级种数）；无可解析届号时返回空列表。
    """
    cohorts = {}
    for record in records:
        match = _COHORT_NUMBER.search(str(record.get("grade", "")))
        if match:
            cohorts[record["name"]] = int(match.group(0))
    if not cohorts:
        return []
    earliest = min(cohorts.values())
    return [r["name"] for r in records if cohorts.get(r["name"]) == earliest]


def _split_major_grade(value) -> tuple[str, str]:
    """把「专业年级」拆成 (专业, 届号)，兼容正序/反序/带空格。"""
    text = _clean(value)
    match = _GRADE_TOKEN.search(text)
    if not match:
        return text, ""
    grade = match.group(0).replace(" ", "").replace("　", "")
    major = (text[:match.start()] + text[match.end():]).strip()
    return major, grade


def _normalize_name(value) -> str:
    """归一化姓名：去掉汉字之间的全角空格（张　爽 -> 张爽）。"""
    text = str(value if value is not None else "")
    text = re.sub(r"(?<=[一-龥])　(?=[一-龥])", "", text)
    return text.strip()


def _clean(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _cell(row, col_idx):
    if col_idx is None or col_idx >= len(row):
        return None
    return row[col_idx].value
